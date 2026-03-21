from pathlib import Path
from datetime import datetime
from urllib.parse import unquote

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from gymapp.models import Client


HEADER_MAP = {
    "ClientID": "client_id",
    "klienti": "client_name",
    "dab. TariRi": "birth_date",
    "piradi nomeri": "passId",
    "sqesi": "gender",
    "telefoni": "phone",
    "el. fosta": "email",
    "organizacia": "organization",
    "baraTi": "card_number",
    "suraTi": "photo",
    "SeniSvna": "comment",
    "UserID": "user_id",
}


GENDER_MAP = {
    "mded": "female",
    "mded.": "female",
    "mde": "female",
    "female": "female",
    "qali": "female",
    "ქალი": "female",
    "მდედ": "female",
    "მდედ.": "female",

    "mamr": "male",
    "mamr.": "male",
    "mamakaci": "male",
    "male": "male",
    "კაცი": "male",
    "მამრ": "male",
    "მამრ.": "male",

    "other": "other",
    "sxva": "other",
    "სხვა": "other",
}


TRANSLIT_SINGLE = {
    "a": "ა", "b": "ბ", "g": "გ", "d": "დ", "e": "ე", "v": "ვ", "z": "ზ",
    "i": "ი", "k": "კ", "l": "ლ", "m": "მ", "n": "ნ", "o": "ო", "p": "პ",
    "r": "რ", "s": "ს", "t": "ტ", "u": "უ", "f": "ფ", "q": "ქ", "y": "ყ",
    "c": "კ", "w": "ვ", "x": "ხ", "h": "ჰ",
    "T": "თ", "S": "შ", "Z": "ძ", "W": "ჭ", "R": "ღ",
}


def translit_to_georgian(text):
    if not text:
        return ""

    s = str(text).strip()
    result = ""

    for ch in s:
        if ch == "'":
            continue
        result += TRANSLIT_SINGLE.get(ch, ch)

    return result


def clean_str(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_indexes(header_row):
    found = {}

    for idx, col_name in enumerate(header_row):
        col_name = str(col_name).strip()
        if col_name in HEADER_MAP:
            found[HEADER_MAP[col_name]] = idx

    return found


def parse_date(value):
    if value in (None, "", " "):
        return None

    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()

    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        return None


def parse_gender(value):
    if not value:
        return None
    raw = str(value).strip().lower()
    return GENDER_MAP.get(raw)


def normalize_georgian_name(full_name):
    full_name = clean_str(full_name)
    if not full_name:
        return "", ""

    georgian_full = translit_to_georgian(full_name)
    parts = georgian_full.split()

    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""

    return parts[0], " ".join(parts[1:])


def get_card_number(raw_value):
    if raw_value in (None, "", " "):
        return ""

    text = str(raw_value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    return text


def get_pass_id(raw_value):
    if raw_value in (None, "", " "):
        return ""

    text = str(raw_value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    return text


def get_photo_db_path(photo_path_raw):
    if not photo_path_raw:
        return ""

    path = unquote(str(photo_path_raw)).replace("\\", "/").strip()

    if not path:
        return ""

    if path.startswith("/media/"):
        path = path[len("/media/"):]

    while path.startswith("../"):
        path = path[3:]

    path = path.lstrip("/")
    return path


class Command(BaseCommand):
    help = "Excel-იდან Client-ების იმპორტი"
    Client.objects.all().delete()

    def add_arguments(self, parser):
        parser.add_argument("excel_file", type=str, help="Excel ფაილის ბილიკი (.xlsx)")
        parser.add_argument(
            "--update-existing",
            action="store_true",
            help="თუ passId ან card_number უკვე არსებობს, Client განაახლოს",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        excel_file = Path(options["excel_file"])
        update_existing = options["update_existing"]

        if not excel_file.exists():
            raise CommandError(f"ფაილი ვერ მოიძებნა: {excel_file}")

        wb = load_workbook(excel_file, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Excel ცარიელია")

        header = rows[0]
        data_rows = rows[1:]

        header_indexes = find_header_indexes(header)

        required = ["client_name", "phone", "passId"]
        missing = [r for r in required if r not in header_indexes]
        if missing:
            raise CommandError(f"სავალდებულო სვეტები ვერ მოიძებნა: {', '.join(missing)}")

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row_num, row in enumerate(data_rows, start=2):
            try:
                full_name = clean_str(row[header_indexes["client_name"]]) if "client_name" in header_indexes else ""
                phone = clean_str(row[header_indexes["phone"]]) if "phone" in header_indexes else ""
                pass_id = get_pass_id(row[header_indexes["passId"]]) if "passId" in header_indexes else ""
                birth_date = parse_date(row[header_indexes["birth_date"]]) if "birth_date" in header_indexes else None
                gender = parse_gender(row[header_indexes["gender"]]) if "gender" in header_indexes else None
                email = clean_str(row[header_indexes["email"]]) if "email" in header_indexes else ""
                organization = translit_to_georgian(clean_str(row[header_indexes["organization"]])) if "organization" in header_indexes else ""
                card_number = get_card_number(row[header_indexes["card_number"]]) if "card_number" in header_indexes else ""
                comment = translit_to_georgian(clean_str(row[header_indexes["comment"]])) if "comment" in header_indexes else ""
                photo_path_raw = clean_str(row[header_indexes["photo"]]) if "photo" in header_indexes else ""

                photo_db_path = get_photo_db_path(photo_path_raw)
                first_name, last_name = normalize_georgian_name(full_name)

                if not first_name and not phone:
                    skipped_count += 1
                    self.stdout.write(self.style.WARNING(f"რიგი {row_num}: გამოტოვებულია"))
                    continue

                if not pass_id:
                    skipped_count += 1
                    self.stdout.write(self.style.WARNING(f"რიგი {row_num}: passId ცარიელია"))
                    continue

                if not last_name:
                    last_name = "-"

                client = Client.objects.filter(passId=pass_id).first()
                existed_before = client is not None

                if client is None and card_number:
                    client = Client.objects.filter(card_number=card_number).first()
                    existed_before = client is not None

                if client and not update_existing:
                    skipped_count += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"რიგი {row_num}: passId/card_number უკვე არსებობს, გამოტოვდა -> {pass_id}"
                        )
                    )
                    continue

                if client is None:
                    client = Client()

                client.passId = pass_id
                client.first_name = first_name
                client.last_name = last_name
                client.birth_date = birth_date
                client.gender = gender
                client.phone = phone
                client.email = email
                client.organization = organization
                client.card_number = card_number
                client.comment = comment
                client.photo = photo_db_path
                client.save()

                if existed_before:
                    updated_count += 1
                else:
                    created_count += 1

            except Exception as e:
                skipped_count += 1
                self.stdout.write(self.style.ERROR(f"რიგი {row_num}: შეცდომა -> {e}"))

        self.stdout.write(
            self.style.SUCCESS(
                f"დასრულდა. დამატდა: {created_count}, განახლდა: {updated_count}, გამოტოვდა: {skipped_count}"
            )
        )